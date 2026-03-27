from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill

# 创建工作簿
wb = Workbook()
ws = wb.active
ws.title = "价格评分测算"

# 设置列宽
ws.column_dimensions['A'].width = 20
ws.column_dimensions['B'].width = 18
ws.column_dimensions['C'].width = 12
ws.column_dimensions['D'].width = 15
ws.column_dimensions['E'].width = 18
ws.column_dimensions['F'].width = 15

# 标题行
headers = ['厂家名称', '不含税总价', '税率(如13%)', '异常低价标记', '评审价格(D₁)', '价格得分']
for col, header in enumerate(headers, 1):
    cell = ws.cell(row=1, column=col, value=header)

# 设置标题样式
header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
header_font = Font(color='FFFFFF', bold=True)
for col in range(1, 7):
    cell = ws.cell(row=1, column=col)
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = Alignment(horizontal='center', vertical='center')

# 预填8行数据区域
for row in range(2, 10):
    # 税率默认13%
    ws.cell(row=row, column=3, value='13%')
    
    # 异常低价判断公式 (低于平均值50%)
    ws.cell(row=row, column=4).value = f'=IF(OR(B{row}="", B{row}=0), "", IF(B{row}<AVERAGE($B$2:$B$9)*0.5, "异常低价", ""))'
    
    # 评审价格公式
    ws.cell(row=row, column=5).value = f'=IF(OR(B{row}="", C{row}=""), "", ROUND(B{row}*(1+((MAX($C$2:$C$9)-C{row})*0.12)), 2))'
    
    # 价格得分公式
    ws.cell(row=row, column=6).value = f'=IF(OR(E{row}="", D{row}="异常低价"), IF(D{row}="异常低价", "否决", ""), LET(基准价,$H$2,系数N,0.95,系数E,1,IF(基准价=0,0,IF(AND(E{row}>=系数N*基准价,E{row}<=基准价),100,IF(E{row}>基准价,ROUND(100-ABS(E{row}-基准价)/基准价*100*系数E,2),ROUND(100-ABS(E{row}-系数N*基准价)/基准价*100*(系数E/2),2))))))'

# 右侧辅助计算区
ws['H1'] = '评审基准价(D)'
ws['H1'].font = Font(bold=True)
ws['H1'].fill = PatternFill(start_color='FFC000', end_color='FFC000', fill_type='solid')
ws['H1'].alignment = Alignment(horizontal='center')

# 评审基准价计算公式 (处理异常低价：参与计算但不打分)
ws['H2'] = '=LET(数据区,E2:E9,标记区,D2:D9,有效数,COUNTA(数据区),IF(有效数<=5,ROUND(AVERAGE(数据区),2),LET(剔异常,FILTER(数据区,标记区<>"异常低价"),剔最值,FILTER(剔异常,(剔异常<>MAX(剔异常))*(剔异常<>MIN(剔异常))),剔最值均值,AVERAGE(剔最值),留用,FILTER(剔最值,ABS(剔最值-剔最值均值)/剔最值均值<=0.2),IF(COUNTA(留用)=0,ROUND(剔最值均值,2),ROUND(AVERAGE(留用),2)))))'
ws['H2'].font = Font(bold=True, color='FF0000')

ws['H3'] = '说明：异常低价厂家参与基准值计算，但价格分为0'
ws['H3'].font = Font(italic=True, color='666666')

# 添加使用说明区域
ws['A12'] = '使用说明：'
ws['A12'].font = Font(bold=True, size=12)

instructions = [
    '1. 在A列填写厂家名称，B列填写不含税总价，C列填写税率（默认13%）',
    '2. D列自动判断是否为异常低价（低于平均值50%标记为异常）',
    '3. E列自动计算评审价格（考虑税率调整系数12%）',
    '4. H2单元格自动计算评审基准价（剔除最高最低及偏差>20%的值）',
    '5. F列自动计算价格得分（异常低价厂家显示"否决"，不参与排名）',
    '6. 可扩展：将第9行后的公式向下复制，支持超过8家厂家'
]

for i, text in enumerate(instructions, 13):
    ws.cell(row=i, column=1, value=text)

# 添加公式说明
ws['A20'] = '公式说明：'
ws['A20'].font = Font(bold=True, size=12)

formulas = [
    '评审价格 D₁ = 不含税总价 × (1 + [max(税率)-本税率] × 12%)',
    '评审基准价 D = 有效数≤5时取平均；>5时剔除1高1低及偏差>20%后取平均',
    '价格得分 F：当 0.95D ≤ D₁ ≤ D 时，F=100；',
    '  D₁>D 时，F=100-|D₁-D|÷D×100×E；D₁<0.95D 时，F=100-|D₁-0.95D|÷D×100×(E÷2)',
    '其中 E=1，N=95%'
]

for i, text in enumerate(formulas, 21):
    ws.cell(row=i, column=1, value=text)

# 添加边框
thin_border = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)

for row in range(1, 10):
    for col in range(1, 7):
        ws.cell(row=row, column=col).border = thin_border

# 保存文件
output_path = '/Users/zhouxinghao/.openclaw/workspace/招标价格评分测算模板.xlsx'
wb.save(output_path)
print(f"✅ Excel文件已生成: {output_path}")
